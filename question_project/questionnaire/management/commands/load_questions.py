import openpyxl
import json
from django.core.management.base import BaseCommand
from questionnaire.models import Question, Rule
from django.core.exceptions import ObjectDoesNotExist

class Command(BaseCommand):
    help = 'Load questions and rules from an Excel file'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the Excel file')

    def handle(self, *args, **kwargs):
        file_path = kwargs['file_path']
        wb = openpyxl.load_workbook(file_path)

        # Load questions
        question_sheet = wb['Questions']
        for row in question_sheet.iter_rows(min_row=2, values_only=True):
            question_id, text, options, data_type = row
            
            # Parse options if the data type is 'choice'
            if data_type == 'choice' and isinstance(options, str):
                try:
                    options = json.loads(options)
                except json.JSONDecodeError:
                    self.stdout.write(self.style.ERROR(f'Invalid JSON format for options in question ID {question_id}. Skipping.'))
                    continue
            else:
                options = None  # No options for non-choice questions

            Question.objects.update_or_create(
                id=question_id,
                defaults={
                    'text': text,
                    'options': options,
                    'data_type': data_type,
                }
            )
        
        # Load rules
        rule_sheet = wb['Rules']
        for row in rule_sheet.iter_rows(min_row=2, values_only=True):
            rule_id, question_id, condition, next_question_id, message = row
            
            try:
                question = Question.objects.get(id=question_id)
            except ObjectDoesNotExist:
                self.stdout.write(self.style.ERROR(f'Question ID {question_id} does not exist. Skipping rule ID {rule_id}.'))
                continue
            
            next_question = None
            if next_question_id:
                try:
                    next_question = Question.objects.get(id=next_question_id)
                except ObjectDoesNotExist:
                    self.stdout.write(self.style.ERROR(f'Next Question ID {next_question_id} does not exist. Skipping rule ID {rule_id}.'))
                    continue

            Rule.objects.update_or_create(
                id=rule_id,
                defaults={
                    'question': question,
                    'condition': condition,
                    'next_question': next_question,
                    'message': message,
                }
            )
        
        self.stdout.write(self.style.SUCCESS('Successfully loaded data from Excel'))
